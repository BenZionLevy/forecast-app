import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from scipy import stats
import io
import time
from openpyxl.utils import get_column_letter

# ייבוא ספריית TradingView
from tvDatafeed import TvDatafeed, Interval

st.set_page_config(page_title="ניתוח קורלציות מקצועי", layout="wide", page_icon="📊")

# ==========================================
# פונקציות עזר קטנות
# ==========================================
def safe_round(val, mult=1.0):
    if pd.isna(val): return None
    return round(float(val) * mult, 2)

# ==========================================
# עיצוב CSS מותאם אישית
# ==========================================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Rubik:wght@300;400;500;600;700&display=swap');
    
    .stApp { background: #f8fafc; }
    html, body, [class*="css"] { font-family: 'Rubik', sans-serif; direction: rtl; }

    .main-header {
        text-align: center; padding: 2rem 0 0.5rem 0; font-size: 2.8rem; font-weight: 800;
        background: linear-gradient(135deg, #0f172a 0%, #2563eb 100%);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text; margin-bottom: 0;
    }
    .sub-header { text-align: center; color: #64748b; font-size: 1.1rem; font-weight: 400; margin-bottom: 2.5rem; }

    div[data-testid="stMetric"] {
        background-color: white; border-radius: 12px; padding: 1rem;
        box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05); border: 1px solid #e2e8f0; transition: transform 0.2s ease-in-out;
    }
    div[data-testid="stMetric"]:hover { transform: translateY(-2px); box-shadow: 0 10px 15px -3px rgba(0,0,0,0.1); }
    
    .stat-badge { display: inline-block; padding: 0.35rem 0.8rem; border-radius: 20px; font-size: 0.95rem; font-weight: 600; margin: 0.15rem; }
    .badge-green { background: #ecfdf5; color: #047857; border: 1px solid #a7f3d0; }
    .badge-red   { background: #fef2f2; color: #b91c1c; border: 1px solid #fecaca; }
    .badge-blue  { background: #eff6ff; color: #1d4ed8; border: 1px solid #bfdbfe; }
    .badge-gray  { background: #f8fafc; color: #475569; border: 1px solid #e2e8f0; }
    
    .section-title {
        font-size: 1.35rem; font-weight: 700; color: #1e293b; margin-top: 2rem; margin-bottom: 1.5rem;
        text-align: right; direction: rtl; display: flex; align-items: center; gap: 0.5rem;
    }
    .section-title::after { content: ""; flex: 1; height: 2px; background: #e2e8f0; margin-right: 15px; border-radius: 2px; }

    .info-box {
        background: white; border-right: 5px solid #3b82f6; border-radius: 10px; padding: 1.5rem;
        color: #334155; direction: rtl; text-align: right; font-size: 1.1rem; line-height: 1.6; margin: 1.5rem 0 2.5rem 0;
        box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05);
    }

    @media (max-width: 768px) {
        .main-header { font-size: 2rem; padding: 1rem 0; }
        .sub-header { font-size: 0.95rem; margin-bottom: 1.5rem; }
        .info-box { font-size: 1rem; padding: 1.2rem; }
        .section-title { font-size: 1.2rem; }
        div[data-testid="stMetric"] { margin-bottom: 0.5rem; }
    }
</style>
""", unsafe_allow_html=True)

st.markdown("<h1 class='main-header'>ניתוח קורלציות מקצועי</h1>", unsafe_allow_html=True)
st.markdown("<p class='sub-header'>מצא קשרים מובהקים בין נכסים פיננסיים (Powered by TradingView)</p>", unsafe_allow_html=True)

# ==========================================
# הגדרות נכסים
# ==========================================
DEFAULT_TICKERS = {
    "לאומי": ("LUMI", "TASE"), 
    "פועלים": ("POLI", "TASE"), 
    "דיסקונט": ("DSCT", "TASE"),
    "מדד ת\"א 35": ("TA35", "TASE"), 
    "מדד ת\"א 125": ("TA125", "TASE"), 
    "מדד בנקים 5": ("TA_BANKS5", "TASE"),
    "S&P 500 ETF": ("SPY", "AMEX"), 
    'נאסד"ק 100 ETF': ("QQQ", "NASDAQ"), 
    "USD/ILS": ("USDILS", "FX_IDC")
}

# סל של ~100 מניות ונכסים עבור אופציית הסורק (מצב 5)
SCANNER_BASKET = {
    "S&P 500 (חוזה עתידי)": ("ES1!", "CME_MINI"),
    "NASDAQ 100 (חוזה עתידי)": ("NQ1!", "CME_MINI"),
    "דאו ג'ונס (חוזה עתידי)": ("YM1!", "CBOT_MINI"),
    "USD/ILS (דולר-שקל)": ("USDILS", "FX_IDC"),
    "מדד ת\"א 35": ("TA35", "TASE"),
    "מדד ת\"א 125": ("TA125", "TASE"),
    "מדד הבנקים": ("TA_BANKS5", "TASE"),
    "מדד נדל\"ן": ("TA_REALESTATE", "TASE"),
    "לאומי": ("LUMI", "TASE"), "פועלים": ("POLI", "TASE"), "דיסקונט": ("DSCT", "TASE"),
    "מזרחי טפחות": ("MZTF", "TASE"), "בינלאומי": ("FIBI", "TASE"), 
    "טבע": ("TEVA", "TASE"), "נייס": ("NICE", "TASE"), "אלביט מערכות": ("ESLT", "TASE"),
    "איי.סי.אל": ("ICL", "TASE"), "קבוצת עזריאלי": ("AZRG", "TASE"),
    "טאואר": ("TSEM", "TASE"), "נובה": ("NVMI", "TASE"), "קמטק": ("CAMT", "TASE"),
    "הפניקס": ("PHOE", "TASE"), "הראל השקעות": ("HREL", "TASE"),
    "מליסרון": ("MLSR", "TASE"), "אלוני חץ": ("ALHE", "TASE"), "מבנה": ("MVNE", "TASE"),
    "אמות": ("AMOT", "TASE"), "בזק": ("BEZQ", "TASE"), "שופרסל": ("SAE", "TASE"),
    "שטראוס": ("STRS", "TASE"), "פז נפט": ("PZOL", "TASE"), "דלק קבוצה": ("DLEKG", "TASE"),
    "אנרג'יאן": ("ENOG", "TASE"), "שפיר הנדסה": ("SPEN", "TASE"), "אל על": ("ELAL", "TASE"),
    "ישראמקו": ("ISRA.L", "TASE"), "אורמת טכנו": ("ORA", "TASE"), "שיכון ובינוי": ("SKBN", "TASE"),
    "אשטרום קבוצה": ("ASHG", "TASE"), "מגדל ביטוח": ("MGDL", "TASE"), "כלל עסקי ביטוח": ("CLIS", "TASE"),
    "דניה סיבוס": ("DANY", "TASE"), "פוקס": ("FOX", "TASE"), "אלקטרה": ("ELTR", "TASE"),
    "תדיראן גרופ": ("TDIR", "TASE"), "איתוראן": ("ITRN", "TASE"), "חילן": ("HLAN", "TASE"),
    "מטריקס": ("MTRX", "TASE"), "מלם תים": ("MLTM", "TASE"), "אודיוקודס": ("AUDC", "TASE"),
    "וואן טכנולוגיות": ("ONE", "TASE"), "דלתא מותגים": ("DLTI", "TASE"), "גב ים": ("GVAM", "TASE"),
    "ריט 1": ("RIT1", "TASE"), "פרשקובסקי": ("PRSK", "TASE"), "אפריקה מגורים": ("AFRE", "TASE"),
    "ישראל קנדה": ("ISCN", "TASE"), "אקרו": ("ACRO", "TASE"), "דיפלומה": ("DIPL", "TASE"),
    "אאורה": ("AURA", "TASE"), "רמי לוי": ("RMLI", "TASE"), "קרסו מוטורס": ("CRSO", "TASE"), 
    "דלק רכב": ("DLEA", "TASE"), "מנורה מבטחים": ("MMHD", "TASE"), "ישראכרט": ("ISCD", "TASE"),
    "מימון ישיר": ("DIFI", "TASE"), "ביג": ("BIG", "TASE"), "אינרום": ("INRM", "TASE"), 
    "אפי נכסים": ("AFPR", "TASE"), "הכשרת הישוב": ("ILDC", "TASE"), "חג'ג'": ("HGG", "TASE"), 
    "י.ח דמרי": ("DIMRI", "TASE"), "נכסים ובנין": ("PTBL", "TASE"), "סלע נדלן": ("SLAR", "TASE"), 
    "קרדן נדלן": ("KRDL", "TASE"), "רוטשטיין": ("ROTS", "TASE"), "רני צים": ("RNZM", "TASE"), 
    "אנלייט אנרגיה": ("ENLT", "TASE"), "אנרג'יקס": ("ENRG", "TASE"), "משק אנרגיה": ("MSKE", "TASE"), 
    "דוראל אנרגיה": ("DORL", "TASE"), "נופר אנרגיה": ("NOFR", "TASE"), "אלקו": ("ELCO", "TASE"), 
    "ארד": ("ARD", "TASE"), "קליל": ("KLIL", "TASE"), "רב בריח": ("RVBR", "TASE"), 
    "קסטרו": ("CAST", "TASE"), "גולף": ("GOLF", "TASE"), "המשביר לצרכן": ("MSBR", "TASE"), 
    "טיב טעם": ("TIVT", "TASE"), "יוחננוף": ("YHNF", "TASE"), "ויקטורי": ("VCTR", "TASE"), 
    "מקס סטוק": ("MAXO", "TASE"), "אלטשולר שחם": ("ALTF", "TASE"), "פריורטק": ("PRTC", "TASE")
}

st.markdown("<div class='section-title'>⚙️ שלב 1: הגדרות הניתוח</div>", unsafe_allow_html=True)

with st.expander("לחץ כאן לפתיחה/סגירה של פאנל ההגדרות", expanded=True):
    col_opt1, col_opt2, col_opt3 = st.columns(3)
    
    with col_opt1:
        st.markdown("**1️⃣ בחירת נכס מטרה**")
        use_custom = st.checkbox("הזן טיקר חופשי (מתקדם)", value=False)
        if use_custom:
            st.caption("פורמט: בורסה:סימול (למשל TASE:LUMI)")
            custom1 = st.text_input("נכס 1 (מטרה):", value="TASE:LUMI").upper().strip()
            custom2 = st.text_input("נכס 2 (להשוואה רגילה):", value="NASDAQ:MSFT").upper().strip()
            asset1_name, asset2_name = custom1, custom2
            try:
                exch1, sym1 = custom1.split(":")
                exch2, sym2 = custom2.split(":")
                ticker1_tuple = (sym1, exch1)
                ticker2_tuple = (sym2, exch2)
            except:
                st.error("⚠️ אנא הקפד על הפורמט: בורסה:סימול (עם נקודתיים באמצע).")
                st.stop()
        else:
            ticker_names = list(DEFAULT_TICKERS.keys())
            asset1_name = st.selectbox("נכס 1 (מטרה)", ticker_names, index=0) # ברירת מחדל לאומי
            asset2_name = st.selectbox("נכס 2 (להשוואה רגילה)", ticker_names, index=3)
            ticker1_tuple = DEFAULT_TICKERS[asset1_name]
            ticker2_tuple = DEFAULT_TICKERS[asset2_name]

    with col_opt2:
        st.markdown("**2️⃣ זמנים וסוג פעולה**")
        mode = st.radio("מבנה הניתוח:", [
            "1. יומי: שער סגירה רשמי", 
            "2. יומי: שעה קבועה ביום",
            "3. מהלך מסחר: חלון שעות", 
            "4. תוך-יומי: קפיצות זמן",
            "5. סורק שוק מורחב (מי מוביל את המניה?)"
        ])
        return_type = st.radio("סוג תשואה:", ["אחוז שינוי רגיל (Simple)", "תשואה לוגריתמית (Log)"])
        use_log_returns = "לוגריתמית" in return_type

    with col_opt3:
        st.markdown("**3️⃣ חלון זמן ומתקדם**")
        start_hour, end_hour, target_hour = None, None, None
        interval_choice, lag_minutes = "1d", 0
        max_lag_to_check = 6 # רלוונטי למצב 5
        
        is_daily_mode = mode == "1. יומי: שער סגירה רשמי"
        max_days = 500 if is_daily_mode else 30 
        default_days = 200 if is_daily_mode else 10

        if mode == "2. יומי: שעה קבועה ביום":
            target_hour = st.selectbox("בחר שעה קבועה:", [f"{h:02d}:00" for h in range(8, 23)], index=2)
            interval_choice = "5m"
        elif mode in ["3. מהלך מסחר: חלון שעות", "4. תוך-יומי: קפיצות זמן"]:
            col_h1, col_h2 = st.columns(2)
            with col_h1: start_hour = st.selectbox("שעת התחלה:", [f"{h:02d}:00" for h in range(8, 23)], index=2)
            with col_h2: end_hour = st.selectbox("שעת סיום:", [f"{h:02d}:00" for h in range(8, 23)], index=8)
            
            if mode == "4. תוך-יומי: קפיצות זמן":
                int_map = {"5 דקות": "5m", "15 דקות": "15m", "30 דקות": "30m", "1 שעה": "60m"}
                interval_choice = int_map[st.selectbox("גודל קפיצה:", list(int_map.keys()))]
                lag_minutes = st.number_input("השהיה לנכס 2 (בדקות):", min_value=0, max_value=600, value=0, step=5)
            else:
                interval_choice = "5m"
        elif mode == "5. סורק שוק מורחב (מי מוביל את המניה?)":
            int_map = {"5 דקות": "5m", "15 דקות": "15m", "30 דקות": "30m", "1 שעה": "60m", "יומי": "1d"}
            interval_choice = int_map[st.selectbox("רזולוציית סריקה:", list(int_map.keys()), index=0)]
            max_lag_to_check = st.number_input("כמה נרות לבדוק אחורה/קדימה (Lag)?", min_value=1, max_value=20, value=6)

        days_back = st.number_input("כמה ימים אחורה לנתח?", min_value=1, max_value=max_days, value=default_days)
    
    st.divider()
    c_adv1, c_adv2 = st.columns(2)
    with c_adv1:
        show_rolling = st.checkbox("הצג מפת קורלציה מתגלגלת (Rolling Correlation)", value=True, disabled=(mode=="5. סורק שוק מורחב (מי מוביל את המניה?)"))
        if show_rolling and mode != "5. סורק שוק מורחב (מי מוביל את המניה?)":
            rolling_window = st.slider("גודל חלון Rolling:", min_value=5, max_value=100, value=20)
    with c_adv2:
        show_ccf = st.checkbox("🔍 מצא מי מגיב למי (Cross-Correlation)", value=False, disabled=(mode=="5. סורק שוק מורחב (מי מוביל את המניה?)"))
        if show_ccf and mode != "5. סורק שוק מורחב (מי מוביל את המניה?)":
            ccf_max_lag = st.slider("מספר השהיות מקסימלי לבדיקה:", min_value=1, max_value=20, value=10)

# ==========================================
# פונקציות חישוב ומשיכה
# ==========================================
@st.cache_data(ttl=600, show_spinner=False)
def fetch_data_tv(sym1_tuple, sym2_tuple, days, interval_str):
    try:
        tv = TvDatafeed()
        tv_intervals = {"1d": Interval.in_daily, "5m": Interval.in_5_minute, "15m": Interval.in_15_minute, "30m": Interval.in_30_minute, "60m": Interval.in_1_hour}
        inter = tv_intervals.get(interval_str, Interval.in_daily)
        bars_per_day = 1 if interval_str == "1d" else (8 * 60) // int(interval_str.replace('m',''))
        total_bars = min(days * bars_per_day, 4900)
        
        df1 = tv.get_hist(symbol=sym1_tuple[0], exchange=sym1_tuple[1], interval=inter, n_bars=total_bars)
        df2 = tv.get_hist(symbol=sym2_tuple[0], exchange=sym2_tuple[1], interval=inter, n_bars=total_bars)
        
        if df1 is None or df1.empty or df2 is None or df2.empty: return pd.DataFrame()
            
        s1 = df1['close'].rename(sym1_tuple[0])
        s2 = df2['close'].rename(sym2_tuple[0])
        
        combined = pd.DataFrame({sym1_tuple[0]: s1, sym2_tuple[0]: s2}).ffill().dropna(how="all")
        
        if combined.index.tz is None: combined.index = combined.index.tz_localize("UTC").tz_convert("Asia/Jerusalem")
        else: combined.index = combined.index.tz_convert("Asia/Jerusalem")
            
        return combined
    except Exception as e:
        return pd.DataFrame()

def compute_stats(s1, s2):
    clean = pd.DataFrame({"a": s1, "b": s2}).dropna()
    if len(clean) < 3: return {"corr": np.nan, "r2": np.nan, "pvalue": np.nan, "n": len(clean)}
    r, p = stats.pearsonr(clean["a"], clean["b"])
    return {"corr": r, "r2": r ** 2, "pvalue": p, "n": len(clean)}

def pvalue_label(p):
    if np.isnan(p): return "—"
    if p < 0.001: return "p < 0.001 ✅"
    if p < 0.01:  return f"p = {p:.3f} ✅"
    if p < 0.05:  return f"p = {p:.3f} ⚠️"
    return f"p = {p:.3f} ❌ לא מובהק"

def calculate_returns(df, is_log):
    if is_log: return np.log(df / df.shift(1))
    return df.pct_change()

# פונקציה לסריקת שוק מרובה נכסים
def run_market_scanner(target_tuple, basket_dict, days, interval_str, max_lags, is_log):
    tv = TvDatafeed()
    tv_intervals = {"1d": Interval.in_daily, "5m": Interval.in_5_minute, "15m": Interval.in_15_minute, "30m": Interval.in_30_minute, "60m": Interval.in_1_hour}
    inter = tv_intervals.get(interval_str, Interval.in_daily)
    bars_per_day = 1 if interval_str == "1d" else (8 * 60) // int(interval_str.replace('m',''))
    total_bars = min(days * bars_per_day, 4900)

    df_target = tv.get_hist(symbol=target_tuple[0], exchange=target_tuple[1], interval=inter, n_bars=total_bars)
    if df_target is None or df_target.empty: return pd.DataFrame()
    s_target = np.log(df_target['close'] / df_target['close'].shift(1)) if is_log else df_target['close'].pct_change()
    s_target = s_target.dropna()

    results = []
    items = list(basket_dict.items())
    
    progress_bar = st.progress(0)
    status_text = st.empty()

    for i, (name, sym_tuple) in enumerate(items):
        # מניעת השוואה של הנכס לעצמו בטבלה
        if sym_tuple[0] == target_tuple[0] and sym_tuple[1] == target_tuple[1]:
            progress_bar.progress((i + 1) / len(items))
            continue
            
        status_text.text(f"🔍 בודק קורלציה מול: {name}...")
        try:
            df_asset = tv.get_hist(symbol=sym_tuple[0], exchange=sym_tuple[1], interval=inter, n_bars=total_bars)
            if df_asset is not None and not df_asset.empty:
                s_asset = np.log(df_asset['close'] / df_asset['close'].shift(1)) if is_log else df_asset['close'].pct_change()
                aligned = pd.DataFrame({"target": s_target, "asset": s_asset}).dropna()
                
                if len(aligned) > 30:
                    best_lag = 0
                    best_corr = 0
                    
                    for lag in range(-max_lags, max_lags + 1):
                        shifted = aligned["asset"].shift(lag)
                        temp = pd.DataFrame({"target": aligned["target"], "asset": shifted}).dropna()
                        if len(temp) > 30:
                            c, _ = stats.pearsonr(temp["target"], temp["asset"])
                            if abs(c) > abs(best_corr):
                                best_corr = c
                                best_lag = lag

                    time_unit = "דקות" if 'm' in interval_str else "ימים"
                    mins = int(interval_str.replace('m','')) if 'm' in interval_str else 1
                    time_diff = abs(best_lag) * mins
                    
                    if best_lag > 0:
                        meaning = f"הנכס מקדים את המניה ב-{time_diff} {time_unit}"
                    elif best_lag < 0:
                        meaning = f"המניה מקדימה את הנכס ב-{time_diff} {time_unit}"
                    else:
                        meaning = "תנועה מסונכרנת (ללא השהיה)"

                    results.append({
                        "נכס השוואה": name,
                        "קורלציה מקסימלית": best_corr,
                        "זמן השהיה (Lag)": best_lag,
                        "משמעות": meaning
                    })
        except: pass
            
        progress_bar.progress((i + 1) / len(items))
        time.sleep(0.1)

    status_text.empty()
    progress_bar.empty()
    
    if not results: return pd.DataFrame()
    res_df = pd.DataFrame(results)
    res_df['R_abs'] = res_df['קורלציה מקסימלית'].abs()
    return res_df.sort_values(by='R_abs', ascending=False).drop(columns=['R_abs']).reset_index(drop=True)

# ==========================================
# ניתוב למצב סורק שוק (מצב 5)
# ==========================================
if mode == "5. סורק שוק מורחב (מי מוביל את המניה?)":
    st.markdown(f"<div class='section-title'>🌍 סורק שוק גלובלי: מי מזיז את {asset1_name}?</div>", unsafe_allow_html=True)
    
    st.info(f"""
    **סקירה נרחבת מול כ-{len(SCANNER_BASKET)} נכסים ומדדים!**
    המערכת סורקת כעת את מניית המטרה שלך מול רשימה של מעל 100 מניות נבחרות (מת"א 125, מדדים גלובליים ודולר-שקל).
    
    ⏳ **זמן משוער:** כדקה. הסריקה מבוצעת בדיוק על פי הטווחים והימים שהוגדרו (רזולוציה: {interval_choice}, ימים אחורה: {days_back}).
    """)
    
    with st.spinner("שואב נתונים ומחשב קורלציות מתקדמות... אנא המתן."):
        scanner_results = run_market_scanner(
            ticker1_tuple, 
            SCANNER_BASKET, 
            days_back, 
            interval_choice, 
            max_lag_to_check, 
            use_log_returns
        )
    
    if not scanner_results.empty:
        def color_corr(val):
            if isinstance(val, float):
                color = '#047857' if val > 0.4 else '#b91c1c' if val < -0.4 else 'black'
                return f'color: {color}; font-weight: bold;'
            return ''
            
        st.dataframe(
            scanner_results.style.map(color_corr, subset=['קורלציה מקסימלית']).format({'קורלציה מקסימלית': '{:.3f}'}),
            use_container_width=True,
            height=600
        )
        
        best_asset = scanner_results.iloc[0]
        st.success(f"🏆 **הנכס המשפיע ביותר על מניית המטרה:** {best_asset['נכס השוואה']} (קורלציה: {best_asset['קורלציה מקסימלית']:.3f}). \n\n**תזמון:** {best_asset['משמעות']}.")
    else:
        st.warning("לא נמצאו מספיק נתונים לחישוב הסריקה. נסה להגדיל את כמות הימים או לבדוק את הטיקר.")
    
    st.stop() # עוצר כאן כדי לא להציג את הגרפים הרגילים של מצבים 1-4

# ==========================================
# עיבוד הנתונים למצבים 1 עד 4
# ==========================================
if ticker1_tuple == ticker2_tuple:
    st.error("⚠️ בחרת את אותו נכס פעמיים. אנא בחר שני נכסים שונים.")
    st.stop()

with st.spinner("🔄 שואב נתוני TradingView בזמן אמת..."):
    raw_df = fetch_data_tv(ticker1_tuple, ticker2_tuple, days_back, interval_choice)

if raw_df.empty:
    st.error("❌ לא ניתן למשוך נתונים מ-TradingView. בדוק את חיבור האינטרנט או שהטיקרים נכונים.")
    st.stop()

sym1_col = ticker1_tuple[0]
sym2_col = ticker2_tuple[0]

scatter_df = pd.DataFrame()
records = []

if mode == "1. יומי: שער סגירה רשמי":
    returns_df_full = calculate_returns(raw_df, use_log_returns)
    scatter_df = returns_df_full.dropna().rename(columns={sym1_col: asset1_name, sym2_col: asset2_name})
    
    for d, row in raw_df.iterrows():
        r1, r2 = returns_df_full.loc[d, sym1_col], returns_df_full.loc[d, sym2_col]
        records.append({
            "תאריך": d.strftime("%d/%m/%Y"),
            f"סגירה {asset1_name}": safe_round(row[sym1_col]), f"תשואה {asset1_name} (%)": safe_round(r1, 100),
            f"סגירה {asset2_name}": safe_round(row[sym2_col]), f"תשואה {asset2_name} (%)": safe_round(r2, 100),
            "הפרש תשואות (%)": safe_round(r1 - r2, 100) if pd.notna(r1) and pd.notna(r2) else None,
        })

elif mode == "2. יומי: שעה קבועה ביום":
    target_end = f"{int(target_hour[:2]):02d}:59"
    hour_df = raw_df.between_time(target_hour, target_end).dropna(how="all")
    if not hour_df.empty:
        hour_df['date_str'] = hour_df.index.date.astype(str)
        daily = hour_df.groupby('date_str').first()
        returns_df_full = calculate_returns(daily, use_log_returns)
        scatter_df = returns_df_full.dropna().rename(columns={sym1_col: asset1_name, sym2_col: asset2_name})
        scatter_df.index = pd.to_datetime(scatter_df.index)
        
        for d_str in daily.index:
            d_obj = pd.to_datetime(d_str)
            r1, r2 = returns_df_full.loc[d_str, sym1_col], returns_df_full.loc[d_str, sym2_col]
            records.append({
                "תאריך": d_obj.strftime("%d/%m/%Y"),
                f"שער {asset1_name}": safe_round(daily.loc[d_str, sym1_col]), f"תשואה {asset1_name} (%)": safe_round(r1, 100),
                f"שער {asset2_name}": safe_round(daily.loc[d_str, sym2_col]), f"תשואה {asset2_name} (%)": safe_round(r2, 100),
                "הפרש תשואות (%)": safe_round(r1 - r2, 100) if pd.notna(r1) and pd.notna(r2) else None,
            })

elif mode == "3. מהלך מסחר: חלון שעות":
    filtered = raw_df.between_time(start_hour, end_hour)
    dates = np.unique(filtered.index.date)
    calc, calc_dates = [], []
    
    for d in dates:
        try: day = filtered.loc[str(d)]
        except KeyError: continue
            
        v1, v2 = day[sym1_col].dropna(), day[sym2_col].dropna()
        if len(v1) >= 2 and len(v2) >= 2:
            if use_log_returns:
                ret1 = np.log(v1.iloc[-1] / v1.iloc[0]) if v1.iloc[0] > 0 else np.nan
                ret2 = np.log(v2.iloc[-1] / v2.iloc[0]) if v2.iloc[0] > 0 else np.nan
            else:
                ret1, ret2 = (v1.iloc[-1] - v1.iloc[0]) / v1.iloc[0], (v2.iloc[-1] - v2.iloc[0]) / v2.iloc[0]
        else:
            ret1, ret2 = np.nan, np.nan
            
        if pd.notna(ret1) and pd.notna(ret2):
            calc.append({asset1_name: ret1, asset2_name: ret2})
            calc_dates.append(pd.to_datetime(d))
            
        records.append({
            "תאריך": d.strftime("%d/%m/%Y"),
            f"פתיחה {asset1_name}": safe_round(v1.iloc[0]) if len(v1)>0 else None, 
            f"סגירה {asset1_name}": safe_round(v1.iloc[-1]) if len(v1)>0 else None,
            f"תשואת חלון {asset1_name} (%)": safe_round(ret1, 100),
            f"פתיחה {asset2_name}": safe_round(v2.iloc[0]) if len(v2)>0 else None, 
            f"סגירה {asset2_name}": safe_round(v2.iloc[-1]) if len(v2)>0 else None,
            f"תשואת חלון {asset2_name} (%)": safe_round(ret2, 100),
            "הפרש תשואות (%)": safe_round(ret1 - ret2, 100) if pd.notna(ret1) and pd.notna(ret2) else None,
        })
        
    scatter_df = pd.DataFrame(calc)
    if not scatter_df.empty: scatter_df.index = calc_dates

elif mode == "4. תוך-יומי: קפיצות זמן":
    filtered = raw_df.between_time(start_hour, end_hour).copy()
    if lag_minutes > 0:
        mins_map = {"5m": 5, "15m": 15, "30m": 30, "60m": 60}
        shift_periods = int(round(lag_minutes / mins_map[interval_choice]))
        if shift_periods > 0: filtered[sym2_col] = filtered[sym2_col].shift(shift_periods)
            
    returns_df_full = calculate_returns(filtered, use_log_returns)
    scatter_df = returns_df_full.dropna().rename(columns={sym1_col: asset1_name, sym2_col: asset2_name})
    
    for d, row in filtered.iterrows():
        r1, r2 = returns_df_full.loc[d, sym1_col], returns_df_full.loc[d, sym2_col]
        if pd.isna(r1) and pd.isna(r2) and pd.isna(row[sym1_col]) and pd.isna(row[sym2_col]): continue
            
        records.append({
            "תאריך ושעה": d.strftime("%d/%m/%Y %H:%M"),
            f"שער {asset1_name}": safe_round(row[sym1_col]), f"תשואה {asset1_name} (%)": safe_round(r1, 100),
            f"שער {asset2_name}": safe_round(row[sym2_col]), f"תשואה {asset2_name} (%)": safe_round(r2, 100),
            "הפרש תשואות (%)": safe_round(r1 - r2, 100) if pd.notna(r1) and pd.notna(r2) else None,
        })

# ==========================================
# שלב 2: תצוגת תוצאות מצבים 1-4
# ==========================================
st.markdown("<div class='section-title'>📊 שלב 2: תוצאות הניתוח</div>", unsafe_allow_html=True)
st.markdown(f"<p class='sub-header' style='margin-bottom: 1rem;'><span dir='ltr'><b>{asset1_name}</b></span> מול <span dir='ltr'><b>{asset2_name}</b></span></p>", unsafe_allow_html=True)

if scatter_df.empty or len(scatter_df) < 3:
    st.warning("⚠️ לא נמצאו מספיק נתונים. נסה להגדיל את כמות הימים או את חלון המסחר.")
    st.stop()

col_a, col_b = scatter_df.columns[0], scatter_df.columns[1]
stats_res = compute_stats(scatter_df[col_a], scatter_df[col_b])

c1, c2, c3, c4 = st.columns(4)
c1.metric("📈 קורלציה (Pearson)", f"{stats_res['corr']:.3f}")
c2.metric("📐 R² (אחוז הסבר שונות)", f"{stats_res['r2']:.3f}")
c3.metric("🔬 מובהקות סטטיסטית", pvalue_label(stats_res['pvalue']))
c4.metric("📋 תצפיות בחישוב", stats_res['n'])

r_val = stats_res["corr"]
strength = "חזקה מאוד" if abs(r_val) >= 0.8 else "חזקה" if abs(r_val) >= 0.6 else "בינונית" if abs(r_val) >= 0.35 else "חלשה"
direction = "חיובית" if r_val > 0 else "שלילית"
badge_class = "badge-green" if r_val > 0.35 else "badge-red" if r_val < -0.35 else "badge-gray"
sig_text = "מובהקת סטטיסטית ✅" if stats_res["pvalue"] < 0.05 else "לא מובהקת סטטיסטית (ייתכן מקרי) ❌"

st.markdown(f"""
<div class='info-box'>
    🧠 <b>מסקנה:</b> נמצאה קורלציה <span class='stat-badge {badge_class}'>{direction} {strength}</span> 
    והיא {sig_text}. <br><br>
    המשמעות היא ש-<b>{stats_res['r2']*100:.1f}%</b> מתנועת התשואות מוסברת על ידי הקשר בין שני הנכסים.
</div>
""", unsafe_allow_html=True)

# ==========================================
# שלב 3: ויזואליזציה (גרפים)
# ==========================================
st.markdown("<div class='section-title'>📉 שלב 3: ויזואליזציה</div>", unsafe_allow_html=True)

if show_ccf and len(scatter_df) > ccf_max_lag * 2:
    lags = list(range(-ccf_max_lag, ccf_max_lag + 1))
    corrs = []
    
    for lag in lags:
        temp_b = scatter_df[col_b].shift(-lag)
        temp_df = pd.DataFrame({"a": scatter_df[col_a], "b": temp_b}).dropna()
        if len(temp_df) > 3:
            c, _ = stats.pearsonr(temp_df["a"], temp_df["b"])
            corrs.append(c)
        else: corrs.append(np.nan)
            
    ccf_df = pd.DataFrame({"השהיה (Lag)": lags, "קורלציה": corrs})
    fig_ccf = px.bar(ccf_df, x="השהיה (Lag)", y="קורלציה", title="מפת הובלה (Cross-Correlation)", template="plotly_white")
    
    max_idx = ccf_df["קורלציה"].idxmax()
    best_lag = ccf_df.loc[max_idx, "השהיה (Lag)"]
    best_corr = ccf_df.loc[max_idx, "קורלציה"]
    colors = ['#3b82f6'] * len(ccf_df)
    colors[ccf_df[ccf_df["השהיה (Lag)"] == best_lag].index[0]] = '#ef4444'
    
    fig_ccf.update_traces(marker_color=colors)
    fig_ccf.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", margin=dict(t=40, b=20), title_x=0.5)
    st.plotly_chart(fig_ccf, use_container_width=True)
    
    lead_text = f"**{col_a}** מוביל את **{col_b}**" if best_lag > 0 else f"**{col_b}** מוביל את **{col_a}**" if best_lag < 0 else "הנכסים מגיבים יחד."
    st.success(f"📌 הקורלציה החזקה ביותר ({best_corr:.3f}) בהשהיה של **{best_lag}**. מסקנה: {lead_text}")

g1, g2 = st.columns([1, 1])
with g1:
    fig_scatter = px.scatter(scatter_df, x=col_a, y=col_b, trendline="ols", labels={col_a: f"תשואה {col_a}", col_b: f"תשואה {col_b}"}, title="פיזור וקו מגמה", template="plotly_white")
    fig_scatter.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", title_x=0.5)
    fig_scatter.update_traces(marker=dict(size=8, opacity=0.7, color="#2563eb", line=dict(width=1, color="DarkSlateGrey")))
    st.plotly_chart(fig_scatter, use_container_width=True)

with g2:
    if show_rolling and len(scatter_df) >= rolling_window:
        rolling_corr = scatter_df[col_a].rolling(rolling_window).corr(scatter_df[col_b])
        fig_roll = go.Figure()
        fig_roll.add_hline(y=0, line_dash="dash", line_color="#cbd5e1")
        fig_roll.add_trace(go.Scatter(x=rolling_corr.index, y=rolling_corr.values, mode="lines", fill="tozeroy", line=dict(color="#10b981", width=2.5), fillcolor="rgba(16, 185, 129, 0.2)"))
        fig_roll.update_layout(title=f"קורלציה מתגלגלת (חלון: {rolling_window})", title_x=0.5, yaxis=dict(range=[-1.1, 1.1]), margin=dict(t=40, b=20, l=10, r=10), template="plotly_white", paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig_roll, use_container_width=True)

# ==========================================
# שלב 4: נתונים גולמיים וייצוא
# ==========================================
st.markdown("<div class='section-title'>📋 שלב 4: נתונים גולמיים וייצוא לאקסל</div>", unsafe_allow_html=True)
summary_df = pd.DataFrame(records)
t1, t2 = st.columns([2, 1])

with t1: st.dataframe(summary_df, use_container_width=True, height=250)
with t2:
    st.markdown("<br>", unsafe_allow_html=True)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        summary_df.to_excel(writer, index=False, sheet_name='Correlation Data')
        worksheet = writer.sheets['Correlation Data']
        try:
            cols = list(summary_df.columns)
            ret1_col, ret2_col = (f"תשואת חלון {asset1_name} (%)", f"תשואת חלון {asset2_name} (%)") if mode == "3. מהלך מסחר: חלון שעות" else (f"תשואה {asset1_name} (%)", f"תשואה {asset2_name} (%)")
            ret1_idx, ret2_idx = cols.index(ret1_col) + 1, cols.index(ret2_col) + 1
            c1_let, c2_let = get_column_letter(ret1_idx), get_column_letter(ret2_idx)
            form_col_let = get_column_letter(len(cols) + 2)
            num_rows = len(summary_df)
            worksheet[f"{form_col_let}1"] = "קורלציה (אקסל חי)"
            worksheet[f"{form_col_let}2"] = f"=CORREL({c1_let}2:{c1_let}{num_rows+1}, {c2_let}2:{c2_let}{num_rows+1})"
        except: pass

    st.download_button(label="📥 הורד נתונים מלאים לאקסל", data=buffer, file_name=f"correlation_{sym1_col}_{sym2_col}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, type="primary")

# ==========================================
# פוטר (קרדיט והערות משפטיות)
# ==========================================
st.divider()
st.markdown("""
<div style='text-align: center; color: #64748b; font-size: 0.85rem; padding-top: 1rem; padding-bottom: 2rem; direction: rtl;'>
    האתר לצורכי מחקר, ועל אחריות המשתמש. השעות המוצגות הן לפי שעון ישראל.<br>
    לשיתופי פעולה ניתן לפנות ליוצר במייל: <a href="mailto:147590@gmail.com" style="color: #3b82f6; text-decoration: none;" dir="ltr">147590@gmail.com</a>
</div>
""", unsafe_allow_html=True)
